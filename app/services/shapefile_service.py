import zipfile
import shutil
from pathlib import Path
from typing import Optional
import pandas as pd
import geopandas as gpd
from openpyxl import load_workbook

from app.core.config import settings
from app.core.exceptions import FileProcessingError


class ShapefileService:
    @staticmethod
    def create_shapefile_for_edit(gdf: gpd.GeoDataFrame, spk_number: str, work_dir: Path) -> Path:
        try:
            shp_path = work_dir / f"{spk_number}_zones.shp"
            gdf.to_file(shp_path, driver='ESRI Shapefile')

            # Create ZIP with shapefile components
            edit_zip = work_dir / "zones_for_edit.zip"
            with zipfile.ZipFile(edit_zip, 'w', zipfile.ZIP_DEFLATED) as z:
                for ext in ['shp', 'shx', 'dbf', 'prj', 'cpg']:
                    p = shp_path.with_suffix(f'.{ext}')
                    if p.exists():
                        z.write(p, p.name)

            return edit_zip

        except Exception as e:
            raise FileProcessingError(f"Shapefile creation failed: {str(e)}")

    @staticmethod
    def process_excel(excel_path: Path, merged_gdf: gpd.GeoDataFrame, spk_number: str, key_id: str) -> pd.DataFrame:
        try:
            # Load and modify Excel
            wb = load_workbook(excel_path)
            sheet = wb['flight record']

            # Copy column L to column A (insert at beginning)
            orig = [c.value for c in sheet['L']]
            sheet.insert_cols(1)
            for i, v in enumerate(orig, start=1):
                sheet.cell(row=i, column=1).value = v

            wb.save(excel_path)

            # Read modified Excel
            df_flight = pd.read_excel(excel_path, sheet_name='flight record', engine='openpyxl')

            # Filter merged GDF to only zones present in flight record
            serial_col = df_flight.columns[0]
            merged_filtered = merged_gdf[
                merged_gdf['Name'].astype(str).isin(df_flight[serial_col].astype(str))
            ].reset_index(drop=True)

            # Build summary DataFrame
            df_summary = pd.DataFrame({'Name': merged_filtered['Name']})

            def lookup(s, idx):
                sub = df_flight[df_flight[df_flight.columns[0]] == s]
                return sub.iloc[0, idx] if not sub.empty else None

            df_summary['TaskAmount'] = df_summary['Name'].map(lambda s: (lookup(s, 6) or 0) * 1000)
            df_summary['StarFlight'] = df_summary['Name'].map(lambda s: str(lookup(s, 1) or '')[:19])
            df_summary['EndFlight'] = df_summary['Name'].map(
                lambda s: (lambda v: str(v)[:11] + str(v)[-8:])(lookup(s, 1))
            )
            df_summary['Capacity'] = 25
            df_summary['SPKNumber'] = spk_number
            df_summary['KeyID'] = key_id

            # Write back to Excel
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as w:
                df_summary.to_excel(w, sheet_name='Sheet1', index=False)

            return merged_filtered

        except Exception as e:
            raise FileProcessingError(f"Excel processing failed: {str(e)}")

    @staticmethod
    def create_final_shapefile(
        gdf: gpd.GeoDataFrame,
        df_summary: pd.DataFrame,
        spk_number: str,
        work_dir: Path
    ) -> Path:
        try:
            # Merge GDF with summary
            gdf_final = gdf.merge(df_summary, on='Name', how='left')

            # Fill nulls in numeric columns
            for col in ("Height", "Route_Spacing", "Task_Flight_Speed"):
                if col in gdf_final.columns:
                    gdf_final[col] = gdf_final[col].ffill().bfill()

            # Truncate column names to 10 characters (shapefile limitation)
            truncate_map = {
                col: col[:10]
                for col in gdf_final.columns
                if col != 'geometry'
            }
            gdf_final = gdf_final.rename(columns=truncate_map)

            # Reorder so geometry is last
            final_cols = list(truncate_map.values()) + ['geometry']
            gdf_final = gdf_final[final_cols]

            # Create output directory
            out_dir = work_dir / "output"
            if out_dir.exists():
                shutil.rmtree(out_dir)
            out_dir.mkdir()

            # Write shapefile
            final_shp = out_dir / f"{spk_number}.shp"
            gdf_final.to_file(final_shp, driver="ESRI Shapefile")

            # Write CPG file for UTF-8 encoding
            with open(out_dir / f"{spk_number}.cpg", 'w', encoding='utf-8') as f:
                f.write('UTF-8')

            # Create ZIP
            zip_out = work_dir / "final_upload.zip"
            if zip_out.exists():
                zip_out.unlink()

            with zipfile.ZipFile(zip_out, 'w', zipfile.ZIP_DEFLATED) as zout:
                for ext in ['shp', 'shx', 'dbf', 'prj', 'cpg']:
                    p = final_shp.with_suffix(f'.{ext}')
                    if p.exists():
                        zout.write(p, p.name)

            return zip_out

        except Exception as e:
            raise FileProcessingError(f"Final shapefile creation failed: {str(e)}")

    @staticmethod
    def load_shapefile_from_zip(zip_path: Path, work_dir: Path) -> gpd.GeoDataFrame:
        try:
            extract_dir = work_dir / "extracted_shp"
            if extract_dir.exists():
                shutil.rmtree(extract_dir)
            extract_dir.mkdir()

            with zipfile.ZipFile(zip_path, 'r') as z:
                z.extractall(extract_dir)

            shp_files = list(extract_dir.glob('*.shp'))
            if not shp_files:
                raise FileProcessingError("No shapefile found in the uploaded ZIP")

            return gpd.read_file(shp_files[0])

        except Exception as e:
            raise FileProcessingError(f"Failed to load shapefile from ZIP: {str(e)}")
