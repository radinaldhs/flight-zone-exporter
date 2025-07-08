import streamlit as st
import zipfile
import shutil
from pathlib import Path
import pandas as pd
import geopandas as gpd
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from shapely.geometry import LineString
import datetime
import requests
import os
import json
import time
from dotenv import load_dotenv

load_dotenv()

# === ArcGIS Config ===
BASE_URL = "https://maps.sinarmasforestry.com/arcgis/rest/services/PreFo/DroneSprayingVendor/FeatureServer/0"
SERVER_URL = "https://maps.sinarmasforestry.com/arcgis/rest/services/PreFo/DroneSprayingVendor/MapServer"
TOKEN_URL = "https://maps.sinarmasforestry.com/portal/sharing/rest/generateToken"

# --- ArcGIS Token Headers (for all token requests) ---
TOKEN_HEADERS = {
    'Content-Type': 'application/x-www-form-urlencoded',
    'Referer': 'https://maps.sinarmasforestry.com/UploadDroneManagements/',
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36',
    'sec-ch-ua': '"Google Chrome";v="137", "Chromium";v="137", "Not/A)Brand";v="24"',
    'sec-ch-ua-platform': '"macOS"',
    'sec-ch-ua-mobile': '?0',
}

def get_final_token():
    session = requests.Session()
    # Step 1: agasha123
    step1 = session.post(TOKEN_URL, headers=TOKEN_HEADERS, data={
        'request': 'getToken',
        'username': os.getenv('GIS_AUTH_USERNAME'),
        'password': os.getenv('GIS_AUTH_PASSWORD'),
        'expiration': '60',
        'referer': 'https://maps.sinarmasforestry.com',
        'f': 'json'
    }).json()
    step1_token = step1.get('token')
    if not step1_token:
        raise RuntimeError("❌ Failed step 1: agasha123 login")

    # Step 2: scoped token for MapServer
    step2 = session.post(TOKEN_URL, headers=TOKEN_HEADERS, data={
        'request': 'getToken',
        'serverUrl': SERVER_URL,
        'token': step1_token,
        'referer': 'https://maps.sinarmasforestry.com',
        'f': 'json'
    }).json()
    scoped_token = step2.get('token')
    if not scoped_token:
        raise RuntimeError("❌ Failed step 2: scoped token")

    # Step 3: fmiseditor
    step3 = session.post(TOKEN_URL, headers=TOKEN_HEADERS, data={
        'request': 'getToken',
        'username': os.getenv('GIS_USERNAME'),
        'password': os.getenv('GIS_PASSWORD'),
        'expiration': '60',
        'referer': 'https://maps.sinarmasforestry.com',
        'f': 'json'
    }).json()
    final_token = step3.get('token')
    if not final_token:
        raise RuntimeError("❌ Failed step 3: final login")

    return final_token

def delete_spk_on_server(spk):
    session = requests.Session()
    token = get_final_token()

    r = session.get(f"{BASE_URL}/query", params={
        'f': 'json',
        'where': f"SPKNumber='{spk}'",
        'outFields': 'OBJECTID',
        'returnGeometry': 'false',
        'token': token
    })
    oids = [f['attributes']['OBJECTID'] for f in r.json().get('features', [])]
    if not oids:
        return f"🔍 No features found for SPK {spk}"

    result = []
    for oid in oids:
        d = session.post(
            f"{BASE_URL}/applyEdits",
            headers=TOKEN_HEADERS,
            data={
                'f': 'json',
                'deletes': str(oid),
                'token': token
            }
        )
        if not d.ok:
            raise RuntimeError(f"Delete failed for OBJECTID {oid}: {d.status_code} {d.text}")
        result.append(d.json())
    return f"✅ Deleted {len(result)} objects for SPK {spk}"

def upload_shapefile_to_server(zip_path: Path):
    session = requests.Session()
    token = get_final_token()

    with open(zip_path, 'rb') as f:
        files = {
            'file': ('final_upload.zip', f, 'application/zip'),
            'token': (None, token)
        }
        resp = session.post(
            "https://maps.sinarmasforestry.com/portal/sharing/rest/content/features/generate",
            params={
                'filetype': 'shapefile',
                'publishParameters': json.dumps({
                    'name': 'UploadedZone_' + OUT_SPK,
                    'maxRecordCount': 1000,
                    'enforceInputFileSizeLimit': True,
                    'enforceOutputJsonSizeLimit': True,
                }),
                'f': 'json',
                'token': token
            },
            files=files
        )
    if resp.ok:
        return resp.json()
    else:
        raise RuntimeError(f"Upload failed: {resp.status_code} {resp.text}")

def post_apply_edits_dynamic(upload_resp: dict):
    import streamlit as st
    token = get_final_token()
    apply_url = f"{BASE_URL}/applyEdits?token={token}"
    headers = {
        "Content-Type": "application/x-www-form-urlencoded"
    }

    # Debug: show top-level keys of upload_resp
    st.write("upload_resp keys:", list(upload_resp.keys()))

    # Determine where features are in the response
    if "featureCollection" in upload_resp:
        layers = upload_resp["featureCollection"].get("layers", [])
        if layers and "featureSet" in layers[0]:
            features = layers[0]["featureSet"].get("features", [])
        else:
            features = []
    elif "featureSet" in upload_resp:
        features = upload_resp["featureSet"].get("features", [])
    else:
        features = []

    adds = []
    for feat in features:
        adds.append({
            "aggregateGeometries": None,
            "geometry": feat["geometry"],
            "symbol": None,
            "attributes": {
                "FlightID": feat["attributes"].get("Name"),
                "DroneID": feat["attributes"].get("Flight_Con"),
                "DroneCapacity": feat["attributes"].get("DroneCapacity", 25),
                "SPKNumber": OUT_SPK,
                "KeyID": OUT_KEYID,
                "StartFlight": feat["attributes"].get("StartFlight", ""),
                "EndFlight": feat["attributes"].get("EndFlight", ""),
                "ProcessDate": int(time.time() * 1000),
                "Height": feat["attributes"].get("Height", 0),
                "Width": feat["attributes"].get("Route_Spac", 0),
                "Speed": feat["attributes"].get("Task_Fligh", 0),
                "TaskArea": feat["attributes"].get("Task_Area", 0),
                "SprayAmount": feat["attributes"].get("Spray_amou", 0),
                "VendorName": "PT SENTRA AGASHA NUSANTARA",
                "UserID": os.getenv('GIS_USERNAME'),
                "CRT_Date": int(time.time() * 1000),
            }
        })

    payload = {
        "f": "json",
        "adds": json.dumps(adds)
    }

    # Log mapped FlightIDs and Names to Streamlit sidebar
    adds_dict = json.loads(payload["adds"])
    st.write("Mapped FlightIDs:", [item["attributes"]["FlightID"] for item in adds_dict])
    st.write("Corresponding Names:", [feat["attributes"].get("Name") for feat in features])

    response = requests.post(apply_url, data=payload, headers=headers)
    try:
        st.info(f"📡 applyEdits POST: {apply_url}")
        st.write("🗂 Payload:", list(payload))
        st.json(response.json())
    except Exception:
        pass
    return response.json()

# --- Page Configuration (must be first Streamlit command) ---
st.set_page_config(page_title="Flight-Zone Selector & Exporter")

# --- Cleanup any previous artifacts on every reload ---
WORK_DIR = Path("working")
if WORK_DIR.exists():
    shutil.rmtree(WORK_DIR)
WORK_DIR.mkdir()

FINAL_ZIP = Path("final_upload.zip")
if FINAL_ZIP.exists():
    FINAL_ZIP.unlink()

EDIT_ZIP = Path("zones_for_edit.zip")
if EDIT_ZIP.exists():
    EDIT_ZIP.unlink()

# --- Sidebar Inputs ---
st.sidebar.header("Upload & Settings")
zip_file = st.sidebar.file_uploader("1. Upload KML ZIP", type="zip")
excel_file = st.sidebar.file_uploader("2. Upload Excel", type=["xlsx","xlsm","xls"])
OUT_SPK = st.sidebar.text_input("3. SPK number")
OUT_KEYID = st.sidebar.text_input("4. KeyID")

# --- Utility: Parse and merge KML layers ---
def parse_kmls(folder: Path) -> gpd.GeoDataFrame:
    records = []
    for kml in folder.rglob('*.kml'):
        root = ET.parse(kml).getroot()
        for pm in root.findall('.//Placemark'):
            name = pm.findtext('name')
            # build a dict of *all* Data tags, replacing spaces with underscores
            props = {
                d.attrib['name'].replace(' ', '_'): d.findtext('value')
                for d in pm.findall('.//Data')
            }
            # parse coords as before
            coords = [
                tuple(map(float, pt.split(',')[:2]))
                for pt in pm.findtext('.//coordinates','').split()
            ]
            rec = {'Name': name, 'geometry': LineString(coords)}
            rec.update(props)
            records.append(rec)

    gdf = gpd.GeoDataFrame(records, crs='EPSG:4326')
    # coerce the columns we know are numbers – anything else stays as a string
    for numcol in ("Height", "Route_Spacing", "Task_Flight_Speed", "Task_Area", "Flight_Time", "Spray_amount"):
        if numcol in gdf.columns:
            gdf[numcol] = pd.to_numeric(gdf[numcol], errors='coerce')
    return gdf


# Move delete button to top-right corner
with st.container():
    delete_col1, delete_col2, reset_col = st.columns([8, 1, 1])
    with delete_col2:
        if OUT_SPK:
            if st.button("🗑️", help="Delete Existing SPK Data", key="delete_button"):
                st.session_state.show_delete_confirm = True
        else:
            st.button("🗑️", help="Enter SPK number to enable delete", key="delete_disabled", disabled=True)
    with reset_col:
        if st.button("❌", help="Clear all inputs and restart", key="reset_button"):
            keys_to_clear = [
                "zip_file", "excel_file", "edited_zip",
                "OUT_SPK", "OUT_KEYID",
                "show_delete_confirm", "delete_button", "reset_button",
                "confirm_delete_button", "cancel_delete_button"
            ]
            for key in keys_to_clear:
                if key in st.session_state:
                    del st.session_state[key]
            st.session_state.clear()
            st.rerun()

# Confirmation modal
if st.session_state.get("show_delete_confirm"):
    st.warning(f"⚠️ Are you sure you want to delete all features for SPK **{OUT_SPK}**?")
    confirm_col1, confirm_col2 = st.columns(2)
    with confirm_col1:
        if st.button("✅ Yes, Delete", key="confirm_delete_button"):
            with st.spinner(f"Deleting data for SPK {OUT_SPK}..."):
                result = delete_spk_on_server(OUT_SPK)
                st.success(result)
                st.session_state.show_delete_confirm = False
    with confirm_col2:
        if st.button("❌ Cancel", key="cancel_delete_button"):
            st.session_state.show_delete_confirm = False


# --- Step 5: Generate Shapefile ZIP for QGIS editing ---
if st.sidebar.button("5. Generate & Download Shapefile ZIP for QGIS Edit"):
    if not (zip_file and excel_file and OUT_SPK and OUT_KEYID):
        st.sidebar.error("Please provide all inputs before generating.")
    else:
        # save and extract KML zip
        ZIP_IN = WORK_DIR / "data.zip"
        with open(ZIP_IN, 'wb') as f:
            f.write(zip_file.getbuffer())
        with zipfile.ZipFile(ZIP_IN, 'r') as zin:
            zin.extractall(WORK_DIR)
        # parse and write shapefile
        merged = parse_kmls(WORK_DIR)
        shp_path = WORK_DIR / f"{OUT_SPK}_zones.shp"
        merged.to_file(shp_path, driver='ESRI Shapefile')
        # zip shapefile components
        edit_zip = WORK_DIR / "zones_for_edit.zip"
        with zipfile.ZipFile(edit_zip, 'w', zipfile.ZIP_DEFLATED) as z:
            for ext in ['shp','shx','dbf','prj','cpg']:
                p = shp_path.with_suffix(f'.{ext}')
                if p.exists():
                    z.write(p, p.name)
        st.sidebar.success("Shapefile ready for QGIS edit. Download below.")
        st.sidebar.download_button(
            "Download shapefile ZIP for QGIS",
            data=open(edit_zip,'rb').read(),
            file_name=edit_zip.name
        )



def handle_final_upload():
    final_path = WORK_DIR / "final_upload.zip"
    if not final_path.exists():
        # Try generating it automatically if inputs are valid
        if not (zip_file and excel_file and OUT_SPK and OUT_KEYID):
            st.sidebar.error("Final upload ZIP not found. Please complete inputs or generate it manually.")
        else:
            ZIP_IN = WORK_DIR / "data.zip"
            with open(ZIP_IN, 'wb') as f:
                f.write(zip_file.getbuffer())
            with zipfile.ZipFile(ZIP_IN, 'r') as zin:
                zin.extractall(WORK_DIR)
            merged = parse_kmls(WORK_DIR)
            process_pipeline(merged)
            final_path = WORK_DIR / "final_upload.zip"

    if final_path.exists():
        with st.spinner("Uploading final shapefile ZIP..."):
            try:
                upload_result = upload_shapefile_to_server(final_path)
                st.sidebar.success("✅ Uploaded successfully.")
                st.json(upload_result)
                post_apply_edits_dynamic(upload_result)
                st.sidebar.success("✅ applyEdits call made.")
            except Exception as e:
                st.sidebar.error(str(e))

if st.sidebar.button("📤 Upload Final ZIP to maps.sinarmasforestry.com"):
    handle_final_upload()

st.markdown("---")

# --- Step 6: Upload Edited Shapefile ZIP or Skip editing ---
st.header("6. Upload Edited Shapefile ZIP (or skip editing)")
edited_zip = st.file_uploader("Upload edited shapefile ZIP", type="zip")

# --- Shared pipeline: Excel edit, join, export, zip ---
def process_pipeline(merged):
    # prepare output folder
    OUT_DIR = WORK_DIR / "output"
    if OUT_DIR.exists():
        shutil.rmtree(OUT_DIR)
    OUT_DIR.mkdir()
    # prepare Excel
    EXCEL_PATH = WORK_DIR / "data.xlsx"
    with open(EXCEL_PATH, 'wb') as f:
        f.write(excel_file.getbuffer())
    wb = load_workbook(EXCEL_PATH)
    sheet = wb['flight record']
    orig = [c.value for c in sheet['L']]
    sheet.insert_cols(1)
    for i, v in enumerate(orig, start=1):
        sheet.cell(row=i, column=1).value = v
    wb.save(EXCEL_PATH)
    df_f = pd.read_excel(EXCEL_PATH, sheet_name='flight record', engine='openpyxl')
    # filter merged to only zones present in flight record
    serial_col = df_f.columns[0]
    merged_filt = merged[merged['Name'].astype(str).isin(df_f[serial_col].astype(str))].reset_index(drop=True)
    # build Sheet1
    df1 = pd.DataFrame({'Name': merged_filt['Name']})
    def lookup(s, idx):
        sub = df_f[df_f[df_f.columns[0]] == s]
        return sub.iloc[0, idx] if not sub.empty else None
    df1['TaskAmount'] = df1['Name'].map(lambda s: (lookup(s,6) or 0) * 1000)
    df1['StarFlight'] = df1['Name'].map(lambda s: str(lookup(s,1) or '')[:19])
    df1['EndFlight']  = df1['Name'].map(lambda s: (lambda v: str(v)[:11] + str(v)[-8:])(lookup(s,1)))
    df1['Capacity']   = 25
    df1['SPKNumber']  = OUT_SPK
    df1['KeyID']      = OUT_KEYID
    with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl', mode='a', if_sheet_exists='replace') as w:
        df1.to_excel(w, sheet_name='Sheet1', index=False)
    # export final shapefile
    gdf = merged_filt.merge(df1, on='Name', how='left')

    # ────────────────────────────────────────────────
    # 1) Fill nulls in Height and Route_Spacing
    #    forward‐fill then back‐fill so leading & trailing NaNs get replaced
    for col in ("Height", "Route_Spacing", "Task_Flight_Speed"):
        if col in gdf.columns:
            gdf[col] = gdf[col].ffill().bfill()
    # ────────────────────────────────────────────────

    final_shp = OUT_DIR / f"{OUT_SPK}.shp"

    truncate = {
        col: col[:10]
        for col in gdf.columns
        if col != 'geometry'
    }
    # rename your in-memory columns
    gdf = gdf.rename(columns=truncate)

    # reorder so geometry is last
    final_cols = list(truncate.values()) + ['geometry']
    gdf = gdf[final_cols]

    # now write—your in-memory names exactly match the on-disk shapefile fields
    gdf.to_file(final_shp, driver="ESRI Shapefile")

    # write CPG
    with open(OUT_DIR / f"{OUT_SPK}.cpg", 'w', encoding='utf-8') as f:
        f.write('UTF-8')
    # zip final components
    ZIP_OUT = WORK_DIR / "final_upload.zip"
    if ZIP_OUT.exists():
        ZIP_OUT.unlink()
    with zipfile.ZipFile(ZIP_OUT, 'w', zipfile.ZIP_DEFLATED) as zout:
        for ext in ['shp','shx','dbf','prj','cpg']:
            p = final_shp.with_suffix(f'.{ext}')
            if p.exists():
                zout.write(p, p.name)
    st.write("▶︎ Columns in the final GDF:", gdf.columns.tolist())
    st.write("**Final Export Data:**")
    st.dataframe(gdf.drop(columns="geometry").head(10))
    # --- Provide download link for final ZIP ---
    st.success("✅ Final ZIP ready for download.")
    st.download_button(
        "Download Final Upload ZIP",
        data=open(ZIP_OUT,'rb').read(),
        file_name=ZIP_OUT.name
    )


# --- Execute pipeline based on user action ---
if edited_zip:
    # extract into fresh folder
    EDIT_DIR = WORK_DIR / "edited_shp"
    if EDIT_DIR.exists():
        shutil.rmtree(EDIT_DIR)
    EDIT_DIR.mkdir()
    temp_zip = EDIT_DIR / "edited.zip"
    with open(temp_zip, 'wb') as f:
        f.write(edited_zip.getbuffer())
    with zipfile.ZipFile(temp_zip, 'r') as zin:
        zin.extractall(EDIT_DIR)
    # load first shapefile found
    shp_list = list(EDIT_DIR.glob('*.shp'))
    if not shp_list:
        st.error(f"No shapefile found in edited folder {EDIT_DIR}")
    else:
        shp_temp = shp_list[0]
        merged = gpd.read_file(shp_temp)
        st.write("**Edited Zones Loaded:**", merged[['Name','Task_Area']])
        process_pipeline(merged)

        # Offer upload option after processing
        if (WORK_DIR / "final_upload.zip").exists():
            if st.button("📤 Upload Final ZIP to maps.sinarmasforestry.com"):
                with st.spinner("Uploading final shapefile ZIP..."):
                    try:
                        upload_result = upload_shapefile_to_server(WORK_DIR / "final_upload.zip")
                        st.success("✅ Uploaded successfully.")
                        st.json(upload_result)
                        # Apply edits dynamically after upload
                        post_apply_edits_dynamic(upload_result)
                        st.sidebar.success("✅ applyEdits call made.")
                    except Exception as e:
                        st.error(str(e))
elif st.sidebar.button("Skip edit and generate final ZIP"):
    if not (zip_file and excel_file and OUT_SPK and OUT_KEYID):
        st.sidebar.error("Please provide all inputs (ZIP, Excel, SPK, KeyID) before skipping edit.")
    else:
        # parse original KMLs
        ZIP_IN = WORK_DIR / "data.zip"
        with open(ZIP_IN, 'wb') as f:
            f.write(zip_file.getbuffer())
        with zipfile.ZipFile(ZIP_IN, 'r') as zin:
            zin.extractall(WORK_DIR)
        merged = parse_kmls(WORK_DIR)
        process_pipeline(merged)

        # Offer upload option after processing
        if (WORK_DIR / "final_upload.zip").exists():
            if st.button("📤 Upload Final ZIP to maps.sinarmasforestry.com"):
                with st.spinner("Uploading final shapefile ZIP..."):
                    try:
                        upload_result = upload_shapefile_to_server(WORK_DIR / "final_upload.zip")
                        st.success("✅ Uploaded successfully.")
                        st.json(upload_result)
                        # Apply edits dynamically after upload
                        post_apply_edits_dynamic(upload_result)
                        st.sidebar.success("✅ applyEdits call made.")
                    except Exception as e:
                        st.error(str(e))

# --- Footer / Watermark (fixed at bottom) ---
year = datetime.date.today().year
footer_html = f"""
<style>
.footer {{
    position: fixed;
    left: 0;
    bottom: 0;
    width: 100%;
    text-align: center;
    color: #888;
    font-size: 0.8rem;
    padding: 0.5rem 0;
    background-color: rgba(255,255,255,0.9);
    z-index: 1000;
}}
</style>
<div class="footer">© {year} Radinal Dewantara Husein</div>
"""
st.markdown(footer_html, unsafe_allow_html=True)